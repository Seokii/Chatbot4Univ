import threading
import json
import pandas as pd
import torch
from pororo import Pororo
from sentence_transformers import util
import numpy as np
import tensorflow as tf
import openpyxl

from config.DatabaseConfig import *
from utils.Database import Database
from utils.BotServer import BotServer
from utils.Preprocess import Preprocess
from models.intent.IntentModel import IntentModel
from models.ner.NerModel import NerModel
from utils.FindAnswer import FindAnswer

# tensorflow gpu 메모리 관련
# tf는 시작시 메모리를 최대로 할당하기 때문에, 0번 GPU를 1GB 메모리만 사용하도록 설정했음.
gpus = tf.config.experimental.list_physical_devices('GPU')
if gpus:
    try:
        tf.config.experimental.set_visible_devices(gpus[0], 'GPU')
        tf.config.experimental.set_virtual_device_configuration(gpus[0],
                        [tf.config.experimental.VirtualDeviceConfiguration(memory_limit=1024)])
    except RuntimeError as e:
        print(e)


# 전처리 객체 생성
p = Preprocess(word2index_dic='/home/irlab/hoseochatbot/train_tools/dict/chatbot_dict.bin',
               userdic='./utils/user_dic.tsv')

# 의도 파악 모델
intent = IntentModel(model_name='/home/irlab/hoseochatbot/models/intent/intent_model.h5', preprocess=p)

# 개체명 인식 모델
ner = NerModel(model_name='/home/irlab/hoseochatbot/models/ner/ner_model.h5', preprocess=p)

# 엑셀 파일 로드 & 임베딩 데이터 로드
df = pd.read_excel('/home/irlab/hoseochatbot/train_tools/qna/train_data.xlsx')
embedding_data = torch.load('/home/irlab/hoseochatbot/train_tools/qna/embedding_data.pt')

# 입력값 임베딩과 유사도 파악 정의
se = Pororo(task="sentence_embedding", lang="ko")
similar = Pororo(task="similarity", lang="ko")

# 학습 데이터 로그 남기기
import logging
Log_Format = logging.Formatter('%(levelname)s %(asctime)s %(message)s')
"""
logging.basicConfig(filename= "logfile.log",    
                    filemode= "w",
                    format= Log_Format,
                    level= logging.ERROR)
"""
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
streamhandler = logging.StreamHandler()
streamhandler.setFormatter(Log_Format)
logger.addHandler(streamhandler)
import datetime
filehandler = logging.FileHandler('/home/irlab/hoseochatbot/logs/logfile_{:%Y%m%d}.log'.format(datetime.datetime.now()),
                                  encoding='utf-8')
filehandler.setFormatter(Log_Format)
logger.addHandler(filehandler)
import logging.handlers

timedfilehandler = logging.handlers.TimedRotatingFileHandler(filename='/home/irlab/hoseochatbot/logs/logfile.log', when='midnight',
                                                             interval=1, encoding='utf-8')
timedfilehandler.setFormatter(Log_Format)
timedfilehandler.suffix = "%Y%m%d"
logger.addHandler(timedfilehandler)


# 입력 문장 임베딩
def return_sim_question(input_sentence):
    # se = Pororo(task="sentence_embedding", lang="ko")
    input_sentence = se(input_sentence)
    input_sentence = torch.tensor(input_sentence)
    return input_sentence

"""
# 엑셀 파일에 학습 질문 추가하기
def add_question_Excel(input):
    wb = openpyxl.load_workbook(filename='train_tools/qna/train_data.xlsx')
    ws = wb.active
    ws = wb['Sheet1']
    ws.cell(row=ws.max_row+1, column=3).value = str(input)
    wb.save(filename='train_tools/qna/train_data.xlsx')
    wb.close()

def add_question_txt(input):
    data = str(input)
    f = open('questionLog.txt', 'a')
    f.write(data)
    f.close()
"""

# ner, intent이후 similarity 비교하여 답변 출력
def chatbot(input, df):
    input = input.strip()
    # 입력 문장 임베딩
    embedding_sentence = return_sim_question(input)

    # 미리 구해진 임베딩 데이터와 현재 임베딩 데이터의 코사인 유사도 추출
    cos_sim = util.pytorch_cos_sim(embedding_sentence, embedding_data)
    # cos_sim = cos_sim.cpu()
    print("가장 높은 코사인 유사도 idx : ", int(np.argmax(cos_sim)))

    # 유사도가 가장 비슷한 질문 인덱스 반환
    best_sim_idx = int(np.argmax(cos_sim))
    print(df['질문(Query)'][best_sim_idx])

    print("선택된 질문과의 유사도: ", similar(input, df['질문(Query)'][best_sim_idx]))
    print("선택된 질문 = " + df['질문(Query)'][best_sim_idx])

    # 유사도 점수 측정
    score = similar(input, df['질문(Query)'][best_sim_idx]) # 유사도 score 산출

    # 유사도가 가장 비슷한 질문에 해당하는 답변 제공
    answer = df['답변(Answer)'][best_sim_idx]
    imageUrl = df['답변 이미지'][best_sim_idx]
    query_intent = df['의도(Intent)'][best_sim_idx]

    """
    else:
        answer = "질문이 정확하지 않거나 답변할 수 없는 질문입니다.\n 수일내로 업데이트 하겠습니다 :("
        imageUrl = None
    """

    return answer, imageUrl, score, query_intent

def to_client(conn, addr):

    # db = params['db']
    try:
        # db.connect() # 디비 연결

        # 데이터 수신
        read = conn.recv(2048) # 수신 데이터가 있을 때까지 블로킹
        print('======================')
        print('Connection from: %s' % str(addr))

        if read is None or not read:
            # 클라이언트 연결이 끊어지거나 오류가 있는 경우
            print('클라이언트 연결 끊어짐')
            exit(0) # 스레드 강제 종료
        
        # json 데이터로 변환
        recv_json_data = json.loads(read.decode())
        print("데이터 수신 : ", recv_json_data)
        query = recv_json_data['Query']
        
        # 의도 파악
        intent_predict = intent.predict_class(query)
        intent_name = intent.labels[intent_predict]

        # 개체명 파악
        ner_predicts = ner.predict(query)
        ner_tags = ner.predict_tags(query)

        """
        # 답변 검색
        try:
            f = FindAnswer(db)
            answer_text, answer_image = f.search(intent_name, ner_tags)
            answer = f.tag_to_word(ner_predicts, answer_text)
            
        except:
            answer = "테스트 버전입니다. 해당 답변을 업데이트하겠습니다. 감사합니다."
            answer_image = None
        """

        # 답변 검색
        #df_mask = (df['의도(Intent)'] == intent_name)
        #answer_df = df[df_mask]
        #print(answer_df)
        try:
            answer, imageUrl, score, query_intent = chatbot(query, df)
        except:
            answer = "chatbot() 함수에서 오류발생..."
            score = 0
            imageUrl = None
            query_intent = None
        print("답변 검색 결과 answer:\n" + answer)
        print("답변 검색 결과 ImageUrl: %s " % imageUrl)
        print("query_intent : %s" % query_intent)
        print("intent_name : %s" % intent_name)
        if (score < 0.70) or (query_intent != intent_name) :
            answer = "질문이 정확하지 않거나 답변할 수 없는 질문입니다.\n 수일내로 업데이트 하겠습니다 :("
            imageUrl = None
            logger.debug(str(query))

        send_json_data_str = {
            "Query": query,
            "Answer": answer,
            "AnswerImageUrl": "["  + str(imageUrl) + "]",
            "Intent": query_intent,
            "NER": str(ner_predicts)
        }

        message = json.dumps(send_json_data_str)
        conn.send(message.encode())

    except Exception as ex:
        print(ex)

    """    
    finally:
        if db is not None: # DB 접속 종료
            db.close()
        conn.close()
    """

if __name__ == '__main__':
    """
    # 질문/답변 학습 디비 연결 객체 생성
    db = Database(
        host=DB_HOST, user=DB_USER, password=DB_PASSWORD, db_name=DB_NAME
    )
    print("DB connect..")
    """

    # 봇 서버 동작
    port = 5050
    listen = 100
    bot = BotServer(port, listen)
    bot.create_sock()
    print("bot start..")
    
    while True:
        conn, addr = bot.ready_for_client()
        """
        params = {
            "db" : db
        }
        """
        
        client = threading.Thread(target=to_client, args=(
            conn,       # 클라이언트 연결 소켓
            addr,       # 클라이언트 연결 주소 정보
            # params      # 스레드 함수 파라미터
        ))
        client.start()  # 스레드 시작

